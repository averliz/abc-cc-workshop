"""Build ocr_tech_scan.xlsx with two tabs from the comparison matrix."""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ocr_tech_scan.xlsx")

HEADERS = ["Tool", "Type", "Deployment", "Primary Strength", "Notable Weakness", "Cost (~2026)", "License / Lock-in", "Best For"]

ROWS = [
    ["GPT-5 / GPT-5.4", "Frontier VLM", "API (OpenAI, Azure)", "Lowest edit distance on OmniDocBench among frontier APIs", "Costly vs open-weight; no native bboxes", "~$1.75 / $14 per 1M tokens", "Proprietary, vendor lock-in", "Hardest documents where cost is secondary"],
    ["Claude Sonnet 4.6 / Opus 4.7", "Frontier VLM", "API (Anthropic, Bedrock, Vertex)", "Lowest hallucination rate on CC-OCR (~0.09%); 1M-token context; high-res vision", "Higher image-token cost on long PDFs", "Sonnet 4.6: $3 / $15 per 1M tokens", "Proprietary", "Agentic doc workflows; compliance-sensitive extraction"],
    ["Gemini 3 Pro / 2.5 Flash", "Frontier VLM", "API (Google, Vertex)", "Top OmniDocBench score among frontier APIs; cheapest Flash tier", "Grounding/bbox via prompting only", "Gemini 2.5 Flash among cheapest vision APIs", "Proprietary", "Long-document, high-volume, price-sensitive"],
    ["Qwen3-VL / Qwen3.5-VL", "Open-weight VLM", "Self-host / hosted API", "Open-weight SOTA; 32 languages; strong on Asian scripts", "235B variant heavy to self-host", "Free (Apache/Qwen license); hosted ~$0.10\u2013$0.30 / 1M tokens", "Apache-2.0 / Qwen License", "Multilingual self-hosted document AI"],
    ["DeepSeek-OCR-2 (3B)", "Document-specialized VLM", "Self-host / hosted", "SOTA small-model on OmniDocBench v1.5 (91.09); runs on single consumer GPU", "Narrower general-VLM reasoning", "Free; hosted ~$0.15 / 1M tokens", "Open weights", "Dense layout / optical-compression use cases"],
    ["PaddleOCR-VL-1.5 (0.9B)", "Document-specialized VLM", "Self-host", "94.5% OmniDocBench v1.5; 109 languages; NaViT dynamic resolution", "Paddle ecosystem friction (improving)", "Free", "Apache-2.0", "Multilingual, self-hosted, GPU-available"],
    ["AWS Textract", "Managed API (hyperscaler)", "Managed (AWS)", "Mature Forms/Tables/Queries; FedRAMP/HIPAA; Layout feature for RAG", "Quality lags VLM-backed APIs on handwriting/complex tables", "$1.50 / 1k pages (Detect); $65 / 1k (AnalyzeDocument)", "Proprietary; AWS lock-in", "Regulated US enterprise on AWS"],
    ["Google Document AI", "Managed API (hyperscaler)", "Managed (GCP)", "Gemini 3-backed Layout Parser; rich processor catalog (invoices, tax, contracts)", "Processor sprawl; narrower data-residency", "Layout Parser $10 / 1k; Form Parser $30 / 1k", "Proprietary; GCP lock-in", "Enterprises already on Vertex AI / BigQuery"],
    ["Azure AI Document Intelligence", "Managed API (hyperscaler)", "Managed (Azure)", "Prebuilt tax/mortgage/ID models; Custom Generative Extraction (schema-in, JSON-out with citations)", "Dense SKU matrix; handwriting behind Mistral", "Read $1.50 / 1k; Prebuilt $10 / 1k; Custom Gen $30 / 1k", "Proprietary; Azure lock-in", "Microsoft/Azure shops; regulated industries"],
    ["Mistral OCR 3", "API specialist (VLM-backed)", "Managed (Mistral, Azure, Bedrock)", "88.9% handwriting, 96.6% tables (vendor); EU-hosted option", "Vendor-reported headline benchmarks", "$2 / 1k pages ($1 with Batch)", "Proprietary but cross-cloud distributed", "High-accuracy at low cost; EU data residency"],
    ["Reducto", "API specialist (agentic)", "Managed", "Hybrid multi-pass OCR + VLM verification; 99.24% on clinical SLAs", "Slower than single-pass; narrower compliance (SOC 2/HIPAA)", "$0.015 / page (~$15 / 1k)", "Proprietary", "Enterprise accuracy-critical (healthcare, financial)"],
    ["LlamaParse v2", "API specialist", "Managed", "Tiered modes (Fast / Agentic Plus); LlamaIndex-native; pinned model versions", "Credit-based pricing complexity", "Fast ~$1.25 / 1k pages; Agentic Plus ~$12+ / 1k", "Proprietary", "RAG/agent pipelines already on LlamaIndex"],
    ["Unstructured.io (API)", "API specialist", "Managed", "25+ file types; element-tree output; RAG-first", "OCR quality = underlying engine", "Fast $1 / 1k; Hi-Res $10 / 1k", "Proprietary", "RAG ingestion over heterogeneous corpora"],
    ["Tesseract 5.5.2", "OSS engine (classic)", "Self-host (CPU)", "100+ languages; runs anywhere; reference baseline", "Weak on layout/tables/handwriting", "Free", "Apache-2.0", "Clean printed text, offline/edge, archival"],
    ["PaddleOCR 3.4.1", "OSS engine (classic + VLM)", "Self-host", "Most active classic OCR; PP-OCRv5 + PaddleOCR-VL unified", "Paddle-framework-leaning historically", "Free", "Apache-2.0", "Self-hosted multilingual OCR at any scale"],
    ["Surya + Marker", "OSS VLM pipeline", "Self-host (GPU preferred)", "PDF\u2192Markdown; tables/math/multi-column", "GPL-3.0 requires commercial license from Datalab", "Free (OSS) / paid (Datalab commercial)", "GPL-3.0", "Academic / technical PDF conversion, non-commercial"],
    ["olmOCR-2 (7B)", "OSS VLM (Ai2)", "Self-host (GPU)", "Qwen2.5-VL fine-tune; 82.4 on olmOCR-Bench; Markdown + LaTeX", "16\u201320 GB VRAM at FP8; English-leaning", "Free", "Apache-2.0", "LLM pretraining-corpus construction, scale OCR"],
    ["MinerU 3.1", "OSS VLM pipeline", "Self-host (GPU preferred)", "95.69 on OmniDocBench v1.6; native DOCX parsing", "GPU-first; smaller English community", "Free", "Apache-based (as of Apr 2026)", "Commercial self-hosted, top-quality pipeline"],
    ["Docling 2.90", "OSS VLM pipeline (IBM)", "Self-host (CPU + GPU)", "MIT-licensed; pluggable VLMs (Granite-Vision); LF governance", "Slightly below MinerU on headline benchmarks", "Free", "MIT", "Commercial-safe enterprise RAG preprocessing"],
    ["Rossum Aurora", "Enterprise IDP", "SaaS / private cloud", "Transactional T-LLM; template-free; direct ERP connectors", "Narrow to AP/PO/order confirmations", "From ~$1,500/mo; enterprise $30k\u2013$150k+", "Proprietary", "AP automation, transactional documents"],
    ["ABBYY Vantage 3.0", "Enterprise IDP", "SaaS / on-prem", "200+ pre-trained skills; deep compliance; 35+ yr heritage", "Heavy deployment; migration friction", "Enterprise quote, mid-six-figure typical", "Proprietary", "Regulated enterprises needing on-prem"],
    ["Hyperscience Hypercell", "Enterprise IDP", "SaaS / on-prem / air-gapped", "ORCA VLM; blocks-and-flows architecture; government-grade", "Premium pricing; long implementation", "$250k\u2013$1M+ annual", "Proprietary", "Government, FSI, air-gapped HITL"],
    ["UiPath Document Understanding", "Enterprise IDP", "SaaS", "Tightest RPA binding; Autopilot for IXP; Generative Validation", "Consumption (AI Unit) pricing can escalate", "AI Unit consumption; 1\u20132 units/page", "Proprietary", "Shops already on UiPath RPA"],
    ["Tungsten TotalAgility", "Enterprise IDP", "SaaS / on-prem", "Kofax + Ephesoft + TotalAgility rollup; wide ERP/RPA footprint", "SKU confusion post-consolidation", "$100k+ annual floor", "Proprietary", "Large bank/government, existing Kofax base"],
    ["Nanonets", "Mid-market IDP", "SaaS", "Low-code workflow builder; fast TTV", "Depth gaps for regulated enterprise", "Consumption-based; $200 signup credits; plans $0\u2013$999/mo", "Proprietary", "SMB/mid-market, rapid deployment"],
    ["Docsumo", "Mid-market IDP", "SaaS", "Lending/insurance focus; transparent pricing", "Lighter integrations than enterprise tier", "Free (100 credits); Growth $299/mo", "Proprietary", "Mid-market FS/insurance"],
    ["Veryfi", "Vertical (receipts)", "SaaS / Mobile SDK", "Sub-5s receipt extraction; mobile SDK", "Narrow to receipts/invoices/checks", "Free tier; Enterprise from $500/mo", "Proprietary", "Expense mgmt, field service"],
    ["Mathpix", "Vertical (STEM)", "Managed API", "Industry-standard math/chem OCR; LaTeX output", "Narrow to scientific docs", "Per-snip / per-page quotes", "Proprietary", "Publishers, EdTech, LLM STEM training data"],
    ["Ocrolus", "Vertical (lending)", "SaaS", "Bank statements, tax returns, fraud detection", "US lending-specific", "Enterprise quote", "Proprietary", "Consumer/SMB lending"],
    ["MyScript", "Vertical (digital ink)", "SDK / on-device", "Stroke-based handwriting in 70+ languages", "Not image OCR \u2014 digital ink only", "OEM licensing", "Proprietary", "Note-taking apps, OEM integrations"],
]

# Category assignments for Tab 2
CATEGORIES = {
    "Frontier VLMs": [
        ("GPT-5 / GPT-5.4", "$$$", "Hardest documents where cost is secondary"),
        ("Claude Sonnet 4.6 / Opus 4.7", "$$$", "Agentic doc workflows; compliance-sensitive"),
        ("Gemini 3 Pro / 2.5 Flash", "$$", "Long-document, high-volume, price-sensitive"),
    ],
    "Open-Weight VLMs": [
        ("Qwen3-VL / Qwen3.5-VL", "Free / $ (hosted)", "Multilingual self-hosted document AI"),
        ("DeepSeek-OCR-2 (3B)", "Free / $ (hosted)", "Dense layout; single consumer GPU"),
        ("PaddleOCR-VL-1.5 (0.9B)", "Free", "Multilingual, self-hosted, GPU-available"),
    ],
    "Managed APIs \u2014 Hyperscalers": [
        ("AWS Textract", "$$\u2013$$$", "Regulated US enterprise on AWS"),
        ("Google Document AI", "$$", "Enterprises on Vertex AI / BigQuery"),
        ("Azure AI Document Intelligence", "$$\u2013$$$", "Microsoft/Azure shops; regulated"),
    ],
    "Managed APIs \u2014 Specialists": [
        ("Mistral OCR 3", "$", "High-accuracy at low cost; EU residency"),
        ("Reducto", "$$", "Enterprise accuracy-critical (healthcare, financial)"),
        ("LlamaParse v2", "$\u2013$$", "RAG/agent pipelines on LlamaIndex"),
        ("Unstructured.io (API)", "$", "RAG ingestion over heterogeneous corpora"),
    ],
    "Open-Source \u2014 Classic Engines": [
        ("Tesseract 5.5.2", "Free", "Clean printed text, offline/edge, archival"),
        ("PaddleOCR 3.4.1", "Free", "Self-hosted multilingual OCR at any scale"),
    ],
    "Open-Source \u2014 VLM Pipelines": [
        ("MinerU 3.1", "Free", "Commercial self-hosted, top-quality pipeline"),
        ("Docling 2.90", "Free", "Commercial-safe enterprise RAG preprocessing"),
        ("olmOCR-2 (7B)", "Free", "LLM pretraining-corpus construction, scale OCR"),
        ("Surya + Marker", "Free (GPL) / Paid", "Academic / technical PDF, non-commercial"),
    ],
    "Enterprise IDP": [
        ("Hyperscience Hypercell", "$$$$", "Government, FSI, air-gapped HITL"),
        ("ABBYY Vantage 3.0", "$$$$", "Regulated enterprises needing on-prem"),
        ("UiPath Document Understanding", "$$$", "Shops already on UiPath RPA"),
        ("Tungsten TotalAgility", "$$$$", "Large bank/government, existing Kofax base"),
        ("Rossum Aurora", "$$$", "AP automation, transactional documents"),
    ],
    "Mid-Market IDP": [
        ("Nanonets", "$", "SMB/mid-market, rapid deployment"),
        ("Docsumo", "$", "Mid-market FS/insurance"),
    ],
    "Vertical Specialists": [
        ("Veryfi", "$\u2013$$", "Expense mgmt, field service"),
        ("Mathpix", "$$", "Publishers, EdTech, LLM STEM data"),
        ("Ocrolus", "$$$", "Consumer/SMB lending"),
        ("MyScript", "OEM", "Note-taking apps, OEM integrations"),
    ],
}

# --- Styles ---
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
header_fill = PatternFill("solid", fgColor="1F3864")
data_font = Font(name="Arial", size=9)
alt_fill = PatternFill("solid", fgColor="F2F7FB")
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
cat_font = Font(name="Arial", bold=True, size=10, color="1F3864")
cat_fill = PatternFill("solid", fgColor="D6E4F0")

wb = Workbook()

# ============================================================
# TAB 1: Comparison Matrix
# ============================================================
ws1 = wb.active
ws1.title = "Comparison Matrix"

ws1.append(HEADERS)
for cell in ws1[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

for i, row in enumerate(ROWS, start=2):
    ws1.append(row)
    for cell in ws1[i]:
        cell.font = data_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = thin_border
        if i % 2 == 0:
            cell.fill = alt_fill

col_widths = [28, 22, 24, 42, 38, 30, 26, 38]
for i, w in enumerate(col_widths, start=1):
    ws1.column_dimensions[get_column_letter(i)].width = w

ws1.freeze_panes = "A2"
ws1.auto_filter.ref = f"A1:H{len(ROWS)+1}"

# ============================================================
# TAB 2: By Category
# ============================================================
ws2 = wb.create_sheet("By Category")

ws2_headers = ["Category", "Tool", "Pricing Tier", "Best-Fit Use Case"]
ws2.append(ws2_headers)
for cell in ws2[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

row_num = 2
for category, tools in CATEGORIES.items():
    ws2.append([category, "", "", ""])
    for cell in ws2[row_num]:
        cell.font = cat_font
        cell.fill = cat_fill
        cell.border = thin_border
    ws2.merge_cells(f"A{row_num}:D{row_num}")
    row_num += 1

    for tool, pricing, use_case in tools:
        ws2.append(["", tool, pricing, use_case])
        for cell in ws2[row_num]:
            cell.font = data_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border
        row_num += 1

ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 30
ws2.column_dimensions["C"].width = 18
ws2.column_dimensions["D"].width = 48
ws2.freeze_panes = "A2"

wb.save(OUT)
print(f"Wrote {OUT} ({os.path.getsize(OUT):,} bytes)")
