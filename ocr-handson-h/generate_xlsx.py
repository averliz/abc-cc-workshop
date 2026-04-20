from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Colors ───────────────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)
CAT_FILL = PatternFill("solid", fgColor="D6E4F0")
CAT_FONT = Font(name="Arial", size=10, bold=True, color="1F4E79")
ALT_FILL = PatternFill("solid", fgColor="F5F7FA")
THIN_BORDER = Border(
    left=Side("thin", color="B4C6E7"),
    right=Side("thin", color="B4C6E7"),
    top=Side("thin", color="B4C6E7"),
    bottom=Side("thin", color="B4C6E7"),
)
WRAP = Alignment(horizontal="left", vertical="center", wrapText=True)
CENTER = Alignment(horizontal="center", vertical="center")

# ══════════════════════════════════════════════════════════════════════════
# TAB 1: Comparison Matrix
# ══════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Comparison Matrix"

headers = ["Tool", "Category", "Accuracy (2026 Bench)", "Cost/1K Pages",
           "Languages", "Best For", "Maturity", "License"]

rows = [
    ["PaddleOCR-VL 1.5", "Open-source VLM", "94.5% (OmniDocBench)", "~$0.09 (self-hosted)", "109", "Complex docs, tables, formulas", "Production", "Apache 2.0"],
    ["Chandra v0.1.0", "Open-source VLM", "83.1% (olmOCR)", "Self-hosted", "90+", "Layout-aware extraction", "Early production", "Open-source"],
    ["Tesseract 5.5", "Open-source engine", "Moderate (clean text)", "Free", "116", "Simple printed text, edge/offline", "Mature/Legacy", "Apache 2.0"],
    ["Mistral OCR v3", "Commercial API", "High (ELO 1409)", "$1.00", "Multi", "Structured docs, handwriting", "Production", "Proprietary API"],
    ["DeepSeek-OCR", "Open-source VLM", "75.7% (olmOCR)", "Self-hosted", "Multi", "On-premises deployment", "Production", "Open-source"],
    ["olmOCR-2", "Open-source", "82.4% (olmOCR)", "Self-hosted", "Multi", "High-throughput PDF conversion", "Production", "Open-source"],
    ["Google Document AI", "Cloud service", "95.8% (general)", "$0.60\u2013$1.50", "200+", "Degraded docs, GCP pipelines", "Mature", "Proprietary"],
    ["AWS Textract", "Cloud service", "94.2% (general)", "$1.50\u2013$65", "100+", "Table extraction, AWS pipelines", "Mature", "Proprietary"],
    ["Azure Doc Intelligence", "Cloud service", "High", "$0.53\u2013$10", "100+", "Custom models, Microsoft shops", "Mature", "Proprietary"],
    ["GPT-5.4", "Proprietary LLM", "Very high", "~$15", "50+", "Complex reasoning over docs", "Mature", "Proprietary API"],
    ["Gemini Flash 2.0", "Proprietary LLM", "High", "~$0.17", "100+", "Cost-effective LLM OCR", "Production", "Proprietary API"],
    ["Qwen3-VL-235B", "Open-source LLM", "Near GPT-5", "Self-hosted", "32 (OCR)", "Full multimodal understanding", "Production", "Open-source"],
    ["ABBYY Vantage", "Enterprise IDP", "Very high", "Enterprise license", "200+", "Multilingual enterprise, compliance", "Mature", "Commercial"],
    ["Nanonets", "IDP platform", "High", "SaaS tiers", "Multi", "Document automation workflows", "Production", "Commercial"],
    ["Rossum", "IDP platform", "Self-improving", "$18K+/year", "Multi", "Zero-template AP automation", "Production", "Commercial"],
    ["Veryfi", "IDP API", "98.7% (invoice)", "API tiers", "38", "Real-time receipt/invoice", "Production", "Commercial"],
    ["Transkribus", "Specialized", "High (trained models)", "Freemium", "100+", "Historical handwriting", "Mature", "Commercial"],
]

col_widths = [22, 18, 22, 20, 12, 34, 16, 16]

for ci, h in enumerate(headers, 1):
    c = ws1.cell(row=1, column=ci, value=h)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = CENTER
    c.border = THIN_BORDER

for ri, row in enumerate(rows, 2):
    for ci, val in enumerate(row, 1):
        c = ws1.cell(row=ri, column=ci, value=val)
        c.font = BODY_FONT
        c.alignment = WRAP
        c.border = THIN_BORDER
        if ri % 2 == 0:
            c.fill = ALT_FILL

for ci, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(ci)].width = w

ws1.freeze_panes = "A2"
ws1.auto_filter.ref = f"A1:H{len(rows)+1}"

# ══════════════════════════════════════════════════════════════════════════
# TAB 2: Tools by Category
# ══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("By Category")

cat_headers = ["Category", "Tool", "Pricing Tier", "Best-Fit Use Case"]
cat_widths = [22, 24, 24, 42]

categories = [
    ("Open-Source OCR", [
        ("PaddleOCR-VL 1.5", "Free (self-hosted, ~$0.09/1K pages)", "Complex documents with tables, formulas, multilingual text"),
        ("Chandra v0.1.0", "Free (self-hosted)", "Layout-aware extraction from multi-column documents"),
        ("Tesseract 5.5", "Free (CPU-only, zero cost)", "Simple printed text; edge/offline deployments"),
        ("DeepSeek-OCR", "Free (self-hosted, GPU required)", "On-premises deployment with no data egress"),
        ("olmOCR-2", "Free (self-hosted)", "High-throughput batch PDF-to-text conversion"),
        ("EasyOCR", "Free (Apache 2.0)", "Quick prototyping; simple text extraction"),
        ("docTR", "Free (Apache 2.0)", "Structured forms; dual TensorFlow/PyTorch pipelines"),
        ("GOT-OCR 2.0", "Free (research model)", "Unified OCR across text, math, charts, sheet music"),
    ]),
    ("Cloud Services", [
        ("Google Document AI", "Pay-per-use ($0.60\u2013$1.50/1K pages)", "Degraded/scanned documents in GCP pipelines"),
        ("AWS Textract", "Pay-per-use ($1.50\u2013$65/1K pages)", "Table extraction in AWS serverless workflows"),
        ("Azure Document Intelligence", "Pay-per-use ($0.53\u2013$10/1K pages)", "Custom models; Microsoft ecosystem; on-prem containers"),
        ("Mistral OCR v3", "Pay-per-use ($1/1K pages)", "High-quality document parsing at low cost; handwriting"),
    ]),
    ("LLM / Multimodal Vision", [
        ("GPT-5.4", "API (~$15/1K pages)", "High-stakes reasoning over complex documents"),
        ("Claude 4 / Sonnet 4.5", "API (comparable to GPT tier)", "Long-document analysis; faithful extraction for legal/finance"),
        ("Gemini Flash 2.0", "API (~$0.17/1K pages)", "Cost-effective bulk LLM OCR at scale"),
        ("Qwen3-VL-235B", "Free (self-hosted, heavy compute)", "GPT-5-class capabilities without API dependency"),
        ("PaddleOCR-VL 7B", "Free (self-hosted, consumer GPU)", "High-volume extraction; outperforms proprietary APIs"),
    ]),
    ("Enterprise IDP", [
        ("ABBYY Vantage", "Enterprise license (custom pricing)", "Multilingual enterprise; 200+ languages; compliance"),
        ("Nanonets", "SaaS tiers (volume-based)", "Document automation; AP, order processing, insurance"),
        ("Rossum", "Enterprise SaaS ($18K+/year)", "Zero-template AP automation with self-improving AI"),
    ]),
    ("Specialized / Vertical", [
        ("Veryfi", "API tiers (per-document)", "Real-time receipt/invoice extraction for fintech apps"),
        ("Mindee", "API tiers (per-call)", "Fast customizable OCR for developers; invoices, IDs"),
        ("Transkribus", "Freemium (credit-based)", "Historical/archival handwriting recognition"),
        ("Microblink", "SDK license (per-scan)", "ID/passport scanning for KYC compliance"),
        ("OCR Studio", "Enterprise license", "Multi-level passport scanning; in-browser OCR"),
    ]),
]

for ci, h in enumerate(cat_headers, 1):
    c = ws2.cell(row=1, column=ci, value=h)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = CENTER
    c.border = THIN_BORDER

row_idx = 2
for cat_name, tools in categories:
    cat_start = row_idx
    for ti, (tool, pricing, use_case) in enumerate(tools):
        ws2.cell(row=row_idx, column=1, value=cat_name if ti == 0 else "")
        ws2.cell(row=row_idx, column=2, value=tool)
        ws2.cell(row=row_idx, column=3, value=pricing)
        ws2.cell(row=row_idx, column=4, value=use_case)
        for ci in range(1, 5):
            c = ws2.cell(row=row_idx, column=ci)
            c.font = BODY_FONT
            c.alignment = WRAP
            c.border = THIN_BORDER
        if ti == 0:
            ws2.cell(row=row_idx, column=1).font = CAT_FONT
            ws2.cell(row=row_idx, column=1).fill = CAT_FILL
        row_idx += 1

    if len(tools) > 1:
        ws2.merge_cells(start_row=cat_start, start_column=1, end_row=cat_start + len(tools) - 1, end_column=1)
        mc = ws2.cell(row=cat_start, column=1)
        mc.alignment = Alignment(horizontal="left", vertical="top", wrapText=True)

    row_idx += 1  # blank row between categories

for ci, w in enumerate(cat_widths, 1):
    ws2.column_dimensions[get_column_letter(ci)].width = w

ws2.freeze_panes = "A2"

# ── Save ─────────────────────────────────────────────────────────────────
out = r"C:\Users\jerem\Documents\github\abc-cc-workshop\ocr-handson-h\ocr_tech_scan.xlsx"
wb.save(out)

import os
print(f"SUCCESS: {out} ({os.path.getsize(out):,} bytes)")
